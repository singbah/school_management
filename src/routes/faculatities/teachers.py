from fastapi import APIRouter, Request, HTTPException, status, File, UploadFile, Form
from fastapi.responses import FileResponse
from datetime import datetime, timedelta
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from bson import ObjectId

from src.routes.config.settings import USERUPLOAD_FOLDER, MAX_LEN, FILE_EXT
from src.routes.config.security import verify_token
from src.database import db

teacher_bp = APIRouter(prefix="/api/teacher")

@teacher_bp.get("/get-students")
async def get_students(request:Request, course_id:str, academy_year:str):
    try:
        token = request.cookies.get("access_token")
        payload = verify_token(token)
        role = payload.get("role")
        if not payload or role not in ['admin', 'staff']:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="unauthorized attempt"
            )
        
        students = []
        cursor = db.student_courses.find({"course_id":course_id,"academy_year":academy_year}).sort("_id", -1)

        async for c in cursor:
            c["_id"] = str(c["_id"])
            students.append(c)
        
        return {"users":students}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )

@teacher_bp.post("/download-grades")
async def download_grades(request:Request, course_id:str, academy_year:str):
    try:
        token = request.cookies
        payload = verify_token(token.get("access_token"))
        staff_id = payload.get("user_id")

        teacher = await db.facultities.find_one({"staff_id":staff_id})
        full_name = f'{teacher["first_name"]} {teacher["last_name"]}'
        
        cursor = db.student_courses.find({"course_id":course_id, "academy_year":academy_year})

        course_info = await db.courses.find_one({"_id":ObjectId(course_id)})
        file_name = f"{course_info['grade_class']}_{course_info['room']}_{course_info['course_name']}.xlsx"

        grade_sheet = []
        async for c in cursor:
            grade_sheet.append(c)

        folder = os.path.join(USERUPLOAD_FOLDER, "grade_sheets")
        os.makedirs(folder, exist_ok=True)
        filename = os.path.join(folder, file_name)
        path_url = f"/grade_sheet/{file_name}"
        
        df = pd.DataFrame({
            "ID ":[s['student_id'] for s in grade_sheet],
            "Name ":[s['student_name'] for s in grade_sheet],
            "Grades ":["" for _ in range(len(grade_sheet))]
        })
        df.to_excel(filename, index=False, startrow=3)

        wb = load_workbook(filename)
        ws = wb.active
        ws.merge_cells('A1:C1')
        ws.merge_cells('A2:E2')
        ws.merge_cells('A3:C3')
        ws['A1'] = "SCHOOL NAME"
        ws['A2'] = f'Academy : {academy_year} - Course: {grade_sheet[0].get("course_name")}'
        ws['A3'] = f"Instructor: {full_name}"
        ws["A1"].font = Font("Calibri", size=18, bold=True)
        ws["A2"].font = Font("Calibri", size=14, bold=True)
        ws["A3"].font = Font("Calibri", size=14, bold=True)
        wb.save(filename)

        await db.grade_sheets.insert_one({
            "course_id":course_id,
            "path_url":path_url,
            "instructor_name":full_name,
            "academy_year":academy_year,
            "filename":file_name
        })

        return {"path_url":path_url}

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )

@teacher_bp.post("/sumbit-grades")
async def submit_grades(requests:Request, grade_sheet:UploadFile = File(...)):
    try:
        token = requests.cookies
        payload = verify_token(token.get("access_token"))
        print(payload)

        if grade_sheet.content_type not in FILE_EXT:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail="Unsupported content"
            )
        if grade_sheet.size > MAX_LEN:
            raise HTTPException(
                status_code=status.HTTP_413_CONTENT_TOO_LARGE,
                detail="file too large"
            )
        df = pd.read_excel(grade_sheet)
        print(df)

        return {"detail":"File receive"}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )